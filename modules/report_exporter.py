"""
Экспорт отчётов аналитики в Excel.

Использует openpyxl для создания Excel-файлов с таблицами и диаграммами.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import Dict

HEADER_FILL = PatternFill(
    start_color="1a73e8", end_color="1a73e8", fill_type="solid"
)
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
HEADER_FONT_BOLD = Font(bold=True)
HEADER_ALIGN_CENTER = Alignment(horizontal="center")


def create_analytics_report(
    metrics: dict, leads_df: pd.DataFrame, deals_df: pd.DataFrame, filepath: str
) -> bool:
    """
    Создаёт полный отчёт аналитики в Excel.

    Args:
        metrics: Словарь с метриками из calculate_metrics()
        leads_df: DataFrame с лидами
        deals_df: DataFrame со сделками
        filepath: Путь для сохранения файла

    Returns:
        True если успешно
    """
    try:
        wb = Workbook()

        # === Лист 1: Сводка ===
        ws_summary = wb.active
        ws_summary.title = "Сводка"

        _create_summary_sheet(ws_summary, metrics)

        # === Лист 2: Лиды по категориям ===
        ws_leads = wb.create_sheet("Лиды по категориям")
        _create_category_sheet(
            ws_leads, metrics["leads_by_category"], metrics["conversion_by_category"]
        )

        # === Лист 3: Сделки по категориям ===
        ws_deals = wb.create_sheet("Сделки по категориям")
        _create_deals_category_sheet(
            ws_deals, metrics["deals_by_category"], metrics["conversion_by_category"]
        )

        # === Лист 4: Менеджеры ===
        ws_managers = wb.create_sheet("Менеджеры")
        _create_managers_sheet(ws_managers, metrics)

        # === Лист 5: Причины отказа ===
        ws_refusals = wb.create_sheet("Причины отказа")
        _create_refusals_sheet(ws_refusals, metrics)

        # === Лист 6: Детализация лидов ===
        ws_leads_detail = wb.create_sheet("Детализация лидов")
        _create_leads_detail_sheet(ws_leads_detail, leads_df)

        # === Лист 7: Детализация сделок ===
        ws_deals_detail = wb.create_sheet("Детализация сделок")
        _create_deals_detail_sheet(ws_deals_detail, deals_df)

        # Сохраняем файл
        wb.save(filepath)
        return True

    except Exception as e:
        print(f"Ошибка создания отчёта: {e}")
        return False


def _create_summary_sheet(ws, metrics: Dict) -> None:
    """Создаёт лист сводки."""

    # Заголовок
    ws["A1"] = "АНАЛИТИКА БИТРИКС24"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = HEADER_ALIGN_CENTER

    # Дата формирования
    ws["A2"] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True)

    # Основные метрики
    metrics_data = [
        ("Всего лидов", metrics["total_leads"]),
        ("Всего сделок", metrics["total_deals"]),
        ("Конверсия", f"{metrics['conversion_rate']:.1f}%"),
    ]

    ws.append([])  # Пустая строка
    ws.append(["Метрика", "Значение"])

    for metric, value in metrics_data:
        ws.append([metric, value])

    # Стили
    for row in ws.iter_rows(min_row=5, max_row=7, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = HEADER_ALIGN_CENTER
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    # Воронка по стадиям лидов
    ws.append([])
    ws.append(["Воронка лидов по стадиям"])
    ws.append(["Стадия", "Количество"])

    for stage, count in metrics.get("lead_stages", {}).items():
        ws.append([stage, count])

    # Воронка по стадиям сделок
    ws.append([])
    ws.append(["Воронка сделок по стадиям"])
    ws.append(["Стадия", "Количество"])

    for stage, count in metrics.get("deal_stages", {}).items():
        ws.append([stage, count])


def _create_category_sheet(
    ws, leads_by_category: Dict, conversion_by_category: Dict
) -> None:
    """Создаёт лист по категориям (лиды)."""

    ws["A1"] = "ЛИДЫ ПО КАТЕГОРИЯМ"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(["Категория", "Лиды", "Сделки", "Конверсия, %"])

    # Заголовки с форматированием
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT_WHITE
        cell.alignment = HEADER_ALIGN_CENTER

    # Данные
    for category in sorted(
        leads_by_category.keys(), key=lambda x: leads_by_category[x], reverse=True
    ):
        leads = leads_by_category.get(category, 0)
        conv_data = conversion_by_category.get(category, {})
        deals = conv_data.get("deals", 0)
        conversion = conv_data.get("conversion", 0)

        ws.append(
            [
                category if category else "Без категории",
                leads,
                deals,
                f"{conversion:.1f}",
            ]
        )

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


def _create_deals_category_sheet(
    ws, deals_by_category: Dict, conversion_by_category: Dict
) -> None:
    """Создаёт лист по категориям (сделки)."""

    ws["A1"] = "СДЕЛКИ ПО КАТЕГОРИЯМ"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(["Категория", "Сделки", "Лиды", "Конверсия, %"])

    # Данные
    for category in sorted(
        deals_by_category.keys(), key=lambda x: deals_by_category[x], reverse=True
    ):
        deals = deals_by_category.get(category, 0)
        conv_data = conversion_by_category.get(category, {})
        leads = conv_data.get("leads", 0)
        conversion = conv_data.get("conversion", 0)

        ws.append(
            [
                category if category else "Без категории",
                deals,
                leads,
                f"{conversion:.1f}",
            ]
        )


def _create_managers_sheet(ws, metrics: Dict) -> None:
    """Создаёт лист по менеджерам."""

    ws["A1"] = "ЭФФЕКТИВНОСТЬ МЕНЕДЖЕРОВ"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])
    ws.append(["Менеджер", "Лиды", "Сделки", "Конверсия, %"])

    for manager, data in sorted(
        metrics.get("conversion_by_manager", {}).items(),
        key=lambda x: x[1]["deals"],
        reverse=True,
    ):
        if manager:
            ws.append(
                [manager, data["leads"], data["deals"], f"{data['conversion']:.1f}"]
            )


def _create_refusals_sheet(ws, metrics: Dict) -> None:
    """Создаёт лист с причинами отказа."""

    ws["A1"] = "ПРИЧИНЫ ОТКАЗА"
    ws["A1"].font = Font(bold=True, size=14)

    # Лиды
    ws.append([])
    ws.append(["Причины отказа (ЛИДЫ)"])
    ws.append(["Причина", "Количество"])

    for reason, count in sorted(
        metrics.get("lead_refusals", {}).items(), key=lambda x: x[1], reverse=True
    ):
        ws.append([reason if reason else "Без причины", count])

    # Сделки
    ws.append([])
    ws.append(["Причины отказа (СДЕЛКИ)"])
    ws.append(["Причина", "Количество"])

    for reason, count in sorted(
        metrics.get("deal_refusals", {}).items(), key=lambda x: x[1], reverse=True
    ):
        ws.append([reason if reason else "Без причины", count])


def _create_leads_detail_sheet(ws, leads_df: pd.DataFrame) -> None:
    """Создаёт лист с детализацией лидов."""

    ws["A1"] = "ДЕТАЛИЗАЦИЯ ЛИДОВ"
    ws["A1"].font = Font(bold=True, size=14)

    # Выбираем нужные колонки
    columns = [
        "ID",
        "Название лида",
        "Категория",
        "Стадия",
        "Ответственный",
        "Рабочий телефон",
        "Мобильный телефон",
        "Источник телефона",
        "Причина отказа",
        "Дата создания",
    ]

    available_cols = [col for col in columns if col in leads_df.columns]

    # Заголовки
    _write_header_row(ws, 3, available_cols, font=HEADER_FONT_BOLD)

    # Данные (только наши лиды)
    our_leads = (
        leads_df[leads_df["Наш лид"]] if "Наш лид" in leads_df.columns else leads_df
    )

    # Векторизованная запись данных вместо iterrows()
    if not our_leads.empty and available_cols:
        # Фильтруем только нужные колонки
        data_to_write = our_leads[available_cols].fillna("")

        # Записываем данные построчно (векторизованно)
        _write_dataframe_rows(ws, data_to_write, start_row=4)


def _create_deals_detail_sheet(ws, deals_df: pd.DataFrame) -> None:
    """Создаёт лист с детализацией сделок."""

    ws["A1"] = "ДЕТАЛИЗАЦИЯ СДЕЛОК"
    ws["A1"].font = Font(bold=True, size=14)

    # Выбираем нужные колонки
    columns = [
        "ID",
        "Название сделки",
        "Категория",
        "Стадия сделки",
        "Ответственный",
        "Сумма",
        "Валюта",
        "Источник телефона",
        "Причина отказа",
        "Дата создания",
    ]

    available_cols = [col for col in columns if col in deals_df.columns]

    # Заголовки
    _write_header_row(ws, 3, available_cols, font=HEADER_FONT_BOLD)

    # Данные (только наши сделки)
    our_deals = (
        deals_df[deals_df["Наша сделка"]]
        if "Наша сделка" in deals_df.columns
        else deals_df
    )

    # Векторизованная запись данных вместо iterrows()
    if not our_deals.empty and available_cols:
        # Фильтруем только нужные колонки
        data_to_write = our_deals[available_cols].fillna("")

        # Записываем данные построчно (векторизованно)
        _write_dataframe_rows(ws, data_to_write, start_row=4)


def _write_header_row(
    ws, row_index: int, headers: list[str], font: Font = HEADER_FONT_WHITE
) -> None:
    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=row_index, column=col_idx, value=col_name)
        cell.font = font
        cell.alignment = HEADER_ALIGN_CENTER
        cell.fill = HEADER_FILL


def _write_dataframe_rows(ws, data: pd.DataFrame, start_row: int) -> None:
    for row_idx, row_data in enumerate(data.itertuples(index=False), start=start_row):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
