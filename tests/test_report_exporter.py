"""
Модульные тесты для modules/report_exporter.py

Тестирует экспорт отчётов аналитики в Excel.
"""

import unittest
from unittest.mock import patch, MagicMock, call
import pandas as pd
from datetime import datetime
from modules.report_exporter import (
    create_analytics_report,
    _create_summary_sheet,
    _create_category_sheet,
    _create_deals_category_sheet,
    _create_managers_sheet,
    _create_refusals_sheet,
    _create_leads_detail_sheet,
    _create_deals_detail_sheet,
    _write_header_row,
    _write_dataframe_rows,
    HEADER_FILL,
    HEADER_FONT_WHITE,
    HEADER_FONT_BOLD,
    HEADER_ALIGN_CENTER,
)


class TestReportExporter(unittest.TestCase):
    """Тесты для модуля report_exporter.py."""

    # ============================================
    # Тесты для create_analytics_report
    # ============================================

    @patch('modules.report_exporter.Workbook')
    def test_create_analytics_report_success(self, mock_workbook):
        """Тестирует успешное создание отчёта."""
        # Arrange
        mock_wb = MagicMock()
        mock_workbook.return_value = mock_wb
        
        metrics = {
            'total_leads': 100,
            'total_deals': 50,
            'conversion_rate': 50.0,
            'leads_by_category': {'Категория 1': 60, 'Категория 2': 40},
            'deals_by_category': {'Категория 1': 30, 'Категория 2': 20},
            'conversion_by_category': {
                'Категория 1': {'leads': 60, 'deals': 30, 'conversion': 50.0},
                'Категория 2': {'leads': 40, 'deals': 20, 'conversion': 50.0},
            },
            'lead_stages': {'Новая': 50, 'В работе': 30, 'Завершена': 20},
            'deal_stages': {'Новая': 25, 'В работе': 15, 'Завершена': 10},
            'conversion_by_manager': {
                'Менеджер 1': {'leads': 50, 'deals': 25, 'conversion': 50.0},
                'Менеджер 2': {'leads': 50, 'deals': 25, 'conversion': 50.0},
            },
            'lead_refusals': {'Не дозвонился': 10, 'Не заинтересован': 5},
            'deal_refusals': {'Передумал': 3, 'Дорого': 2},
        }
        
        leads_df = pd.DataFrame({
            'ID': [1, 2],
            'Название лида': ['Лид 1', 'Лид 2'],
            'Категория': ['Категория 1', 'Категория 2'],
            'Наш лид': [True, True],
        })
        
        deals_df = pd.DataFrame({
            'ID': [1],
            'Название сделки': ['Сделка 1'],
            'Категория': ['Категория 1'],
            'Наша сделка': [True],
        })
        
        # Act
        result = create_analytics_report(metrics, leads_df, deals_df, 'output.xlsx')
        
        # Assert
        self.assertTrue(result)
        mock_workbook.assert_called_once()
        mock_wb.save.assert_called_once_with('output.xlsx')

    @patch('modules.report_exporter.Workbook')
    def test_create_analytics_report_invalid_path(self, mock_workbook):
        """Тестирует неверный путь к файлу."""
        # Arrange
        mock_workbook.side_effect = IOError('Invalid path')
        
        metrics = {'total_leads': 0, 'total_deals': 0, 'conversion_rate': 0.0}
        leads_df = pd.DataFrame({'ID': [1]})
        deals_df = pd.DataFrame({'ID': [1]})
        
        # Act
        result = create_analytics_report(metrics, leads_df, deals_df, '/invalid/path/output.xlsx')
        
        # Assert
        self.assertFalse(result)

    @patch('modules.report_exporter.Workbook')
    def test_create_analytics_report_creates_all_sheets(self, mock_workbook):
        """Тестирует создание всех листов отчёта."""
        # Arrange
        mock_wb = MagicMock()
        mock_workbook.return_value = mock_wb
        
        metrics = {
            'total_leads': 10,
            'total_deals': 5,
            'conversion_rate': 50.0,
            'leads_by_category': {},
            'deals_by_category': {},
            'conversion_by_category': {},
            'lead_stages': {},
            'deal_stages': {},
            'conversion_by_manager': {},
            'lead_refusals': {},
            'deal_refusals': {},
        }
        leads_df = pd.DataFrame()
        deals_df = pd.DataFrame()
        
        # Act
        create_analytics_report(metrics, leads_df, deals_df, 'output.xlsx')
        
        # Assert
        # Проверка создания всех листов
        self.assertEqual(mock_wb.create_sheet.call_count, 6)
        expected_sheets = [
            "Лиды по категориям",
            "Сделки по категориям",
            "Менеджеры",
            "Причины отказа",
            "Детализация лидов",
            "Детализация сделок",
        ]
        for sheet_name in expected_sheets:
            mock_wb.create_sheet.assert_any_call(sheet_name)

    # ============================================
    # Тесты для _create_summary_sheet
    # ============================================

    def test_create_summary_sheet(self):
        """Тестирует создание сводного листа."""
        # Arrange
        mock_ws = MagicMock()
        
        metrics = {
            'total_leads': 100,
            'total_deals': 50,
            'conversion_rate': 50.0,
            'lead_stages': {'Новая': 50, 'В работе': 30, 'Завершена': 20},
            'deal_stages': {'Новая': 25, 'В работе': 15, 'Завершена': 10},
        }
        
        # Act
        _create_summary_sheet(mock_ws, metrics)
        
        # Assert
        # Проверка что устанавливался заголовок A1
        mock_ws.__setitem__.assert_any_call("A1", "АНАЛИТИКА БИТРИКС24")
        
        # Проверка что устанавливалась дата в A2 (начинается с "Дата формирования:")
        setitem_calls = mock_ws.__setitem__.call_args_list
        a2_call = [call for call in setitem_calls if call[0][0] == "A2"]
        self.assertTrue(len(a2_call) > 0)
        self.assertTrue(str(a2_call[0][0][1]).startswith("Дата формирования:"))
        
        # Проверка вызова append для метрик
        self.assertGreater(mock_ws.append.call_count, 0)

    def test_create_summary_sheet_empty_metrics(self):
        """Тестирует создание сводки с пустыми метриками."""
        # Arrange
        mock_ws = MagicMock()
        
        metrics = {
            'total_leads': 0,
            'total_deals': 0,
            'conversion_rate': 0.0,
            'lead_stages': {},
            'deal_stages': {},
        }
        
        # Act
        _create_summary_sheet(mock_ws, metrics)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "АНАЛИТИКА БИТРИКС24")
        self.assertGreater(mock_ws.append.call_count, 0)

    # ============================================
    # Тесты для _create_category_sheet
    # ============================================

    def test_create_category_sheet(self):
        """Тестирует создание листа категорий лидов."""
        # Arrange
        mock_ws = MagicMock()
        
        leads_by_category = {'Категория A': 100, 'Категория B': 50}
        conversion_by_category = {
            'Категория A': {'leads': 100, 'deals': 50, 'conversion': 50.0},
            'Категория B': {'leads': 50, 'deals': 10, 'conversion': 20.0},
        }
        
        # Act
        _create_category_sheet(mock_ws, leads_by_category, conversion_by_category)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "ЛИДЫ ПО КАТЕГОРИЯМ")
        self.assertGreater(mock_ws.append.call_count, 0)

    def test_create_category_sheet_empty(self):
        """Тестирует создание листа с пустыми категориями."""
        # Arrange
        mock_ws = MagicMock()
        
        # Act
        _create_category_sheet(mock_ws, {}, {})
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "ЛИДЫ ПО КАТЕГОРИЯМ")

    # ============================================
    # Тесты для _create_deals_category_sheet
    # ============================================

    def test_create_deals_category_sheet(self):
        """Тестирует создание листа категорий сделок."""
        # Arrange
        mock_ws = MagicMock()
        
        deals_by_category = {'Категория X': 30, 'Категория Y': 20}
        conversion_by_category = {
            'Категория X': {'leads': 60, 'deals': 30, 'conversion': 50.0},
            'Категория Y': {'leads': 40, 'deals': 20, 'conversion': 50.0},
        }
        
        # Act
        _create_deals_category_sheet(mock_ws, deals_by_category, conversion_by_category)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "СДЕЛКИ ПО КАТЕГОРИЯМ")
        self.assertGreater(mock_ws.append.call_count, 0)

    # ============================================
    # Тесты для _create_managers_sheet
    # ============================================

    def test_create_managers_sheet(self):
        """Тестирует создание листа менеджеров."""
        # Arrange
        mock_ws = MagicMock()
        
        metrics = {
            'conversion_by_manager': {
                'Менеджер 1': {'leads': 100, 'deals': 50, 'conversion': 50.0},
                'Менеджер 2': {'leads': 80, 'deals': 40, 'conversion': 50.0},
            }
        }
        
        # Act
        _create_managers_sheet(mock_ws, metrics)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "ЭФФЕКТИВНОСТЬ МЕНЕДЖЕРОВ")
        self.assertGreater(mock_ws.append.call_count, 0)

    def test_create_managers_sheet_empty(self):
        """Тестирует создание листа менеджеров с пустыми данными."""
        # Arrange
        mock_ws = MagicMock()
        
        metrics = {'conversion_by_manager': {}}
        
        # Act
        _create_managers_sheet(mock_ws, metrics)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "ЭФФЕКТИВНОСТЬ МЕНЕДЖЕРОВ")

    # ============================================
    # Тесты для _create_refusals_sheet
    # ============================================

    def test_create_refusals_sheet(self):
        """Тестирует создание листа причин отказа."""
        # Arrange
        mock_ws = MagicMock()
        
        metrics = {
            'lead_refusals': {'Не дозвонился': 10, 'Не заинтересован': 5},
            'deal_refusals': {'Передумал': 3, 'Дорого': 2},
        }
        
        # Act
        _create_refusals_sheet(mock_ws, metrics)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "ПРИЧИНЫ ОТКАЗА")
        self.assertGreater(mock_ws.append.call_count, 0)

    # ============================================
    # Тесты для _create_leads_detail_sheet
    # ============================================

    def test_create_leads_detail_sheet(self):
        """Тестирует создание листа детализации лидов."""
        # Arrange
        mock_ws = MagicMock()
        
        leads_df = pd.DataFrame({
            'ID': [1, 2, 3],
            'Название лида': ['Лид 1', 'Лид 2', 'Лид 3'],
            'Категория': ['A', 'B', 'C'],
            'Стадия': ['Новая', 'В работе', 'Завершена'],
            'Ответственный': ['Менеджер 1', 'Менеджер 2', 'Менеджер 1'],
            'Рабочий телефон': ['+70000000001', '+70000000002', '+70000000003'],
            'Мобильный телефон': ['', '+70000000004', ''],
            'Источник телефона': ['Яндекс', '2GIS', 'Яндекс'],
            'Причина отказа': ['', 'Не дозвонился', ''],
            'Дата создания': ['2024-01-01', '2024-01-02', '2024-01-03'],
            'Наш лид': [True, True, False],
        })
        
        # Act
        _create_leads_detail_sheet(mock_ws, leads_df)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "ДЕТАЛИЗАЦИЯ ЛИДОВ")
        mock_ws.cell.assert_called()

    def test_create_leads_detail_sheet_empty(self):
        """Тестирует создание листа с пустым DataFrame."""
        # Arrange
        mock_ws = MagicMock()
        leads_df = pd.DataFrame()
        
        # Act
        _create_leads_detail_sheet(mock_ws, leads_df)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "ДЕТАЛИЗАЦИЯ ЛИДОВ")

    # ============================================
    # Тесты для _create_deals_detail_sheet
    # ============================================

    def test_create_deals_detail_sheet(self):
        """Тестирует создание листа детализации сделок."""
        # Arrange
        mock_ws = MagicMock()
        
        deals_df = pd.DataFrame({
            'ID': [1, 2],
            'Название сделки': ['Сделка 1', 'Сделка 2'],
            'Категория': ['A', 'B'],
            'Стадия сделки': ['Новая', 'В работе'],
            'Ответственный': ['Менеджер 1', 'Менеджер 2'],
            'Сумма': [100000, 200000],
            'Валюта': ['RUB', 'RUB'],
            'Источник телефона': ['Яндекс', '2GIS'],
            'Причина отказа': ['', 'Передумал'],
            'Дата создания': ['2024-01-01', '2024-01-02'],
            'Наша сделка': [True, True],
        })
        
        # Act
        _create_deals_detail_sheet(mock_ws, deals_df)
        
        # Assert
        mock_ws.__setitem__.assert_any_call("A1", "ДЕТАЛИЗАЦИЯ СДЕЛОК")
        mock_ws.cell.assert_called()

    # ============================================
    # Тесты для _write_header_row
    # ============================================

    def test_write_header_row(self):
        """Тестирует запись заголовка таблицы."""
        # Arrange
        mock_ws = MagicMock()
        headers = ['ID', 'Название', 'Менеджер', 'Сумма']
        row_index = 1
        
        # Act
        _write_header_row(mock_ws, row_index, headers, font=HEADER_FONT_WHITE)
        
        # Assert
        self.assertEqual(mock_ws.cell.call_count, len(headers))
        
        # Проверка первой ячейки
        mock_ws.cell.assert_any_call(row=1, column=1, value='ID')
        mock_ws.cell.assert_any_call(row=1, column=4, value='Сумма')

    def test_write_header_row_custom_row(self):
        """Тестирует запись заголовка на кастомной строке."""
        # Arrange
        mock_ws = MagicMock()
        headers = ['Колонка 1', 'Колонка 2']
        row_index = 5
        
        # Act
        _write_header_row(mock_ws, row_index, headers)
        
        # Assert
        mock_ws.cell.assert_any_call(row=5, column=1, value='Колонка 1')
        mock_ws.cell.assert_any_call(row=5, column=2, value='Колонка 2')

    def test_write_header_row_empty_headers(self):
        """Тестирует запись заголовка с пустым списком."""
        # Arrange
        mock_ws = MagicMock()
        headers = []
        
        # Act
        _write_header_row(mock_ws, 1, headers)
        
        # Assert
        mock_ws.cell.assert_not_called()

    # ============================================
    # Тесты для _write_dataframe_rows
    # ============================================

    def test_write_dataframe_rows(self):
        """Тестирует запись данных DataFrame."""
        # Arrange
        mock_ws = MagicMock()
        
        df = pd.DataFrame({
            'ID': [1, 2, 3],
            'Название': ['A', 'B', 'C'],
            'Значение': [100, 200, 300],
        })
        
        start_row = 2
        
        # Act
        _write_dataframe_rows(mock_ws, df, start_row)
        
        # Assert
        # Проверка что вызывалось для каждой строки (3 строки * 3 колонки = 9 вызовов)
        self.assertEqual(mock_ws.cell.call_count, 9)
        
        # Проверка первой ячейки
        mock_ws.cell.assert_any_call(row=2, column=1, value=1)
        mock_ws.cell.assert_any_call(row=2, column=2, value='A')

    def test_write_dataframe_rows_with_none(self):
        """Тестирует запись DataFrame с None значениями."""
        # Arrange
        mock_ws = MagicMock()
        
        df = pd.DataFrame({
            'ID': [1, None],
            'Название': [None, 'B'],
        })
        
        # Act
        _write_dataframe_rows(mock_ws, df, start_row=1)
        
        # Assert
        # None должны быть заменены на пустую строку
        calls = mock_ws.cell.call_args_list
        self.assertTrue(any(call[1]['value'] == '' for call in calls))

    def test_write_dataframe_rows_empty(self):
        """Тестирует запись пустого DataFrame."""
        # Arrange
        mock_ws = MagicMock()
        df = pd.DataFrame()
        
        # Act
        _write_dataframe_rows(mock_ws, df, start_row=1)
        
        # Assert
        mock_ws.cell.assert_not_called()

    # ============================================
    # Тесты для констант стилей
    # ============================================

    def test_header_fill_is_pattern_fill(self):
        """Тестирует что HEADER_FILL это PatternFill."""
        from openpyxl.styles import PatternFill
        self.assertIsInstance(HEADER_FILL, PatternFill)

    def test_header_font_white_is_font(self):
        """Тестирует что HEADER_FONT_WHITE это Font."""
        from openpyxl.styles import Font
        self.assertIsInstance(HEADER_FONT_WHITE, Font)
        self.assertTrue(HEADER_FONT_WHITE.bold)
        # Цвет может быть объектом Color, проверяем что он установлен
        self.assertIsNotNone(HEADER_FONT_WHITE.color)

    def test_header_align_center_is_alignment(self):
        """Тестирует что HEADER_ALIGN_CENTER это Alignment."""
        from openpyxl.styles import Alignment
        self.assertIsInstance(HEADER_ALIGN_CENTER, Alignment)
        self.assertEqual(HEADER_ALIGN_CENTER.horizontal, "center")


class TestReportExporterIntegration(unittest.TestCase):
    """Интеграционные тесты для report_exporter.py."""

    @patch('modules.report_exporter.Workbook')
    def test_full_report_with_real_data(self, mock_workbook):
        """Тестирует полный отчёт с реалистичными данными."""
        # Arrange
        mock_wb = MagicMock()
        mock_workbook.return_value = mock_wb
        
        # Реалистичные метрики
        metrics = {
            'total_leads': 250,
            'total_deals': 125,
            'conversion_rate': 50.0,
            'leads_by_category': {
                'Автосервис': 100,
                'Шиномонтаж': 80,
                'Детейлинг': 70,
            },
            'deals_by_category': {
                'Автосервис': 50,
                'Шиномонтаж': 40,
                'Детейлинг': 35,
            },
            'conversion_by_category': {
                'Автосервис': {'leads': 100, 'deals': 50, 'conversion': 50.0},
                'Шиномонтаж': {'leads': 80, 'deals': 40, 'conversion': 50.0},
                'Детейлинг': {'leads': 70, 'deals': 35, 'conversion': 50.0},
            },
            'lead_stages': {
                'Новая': 100,
                'В работе': 80,
                'Завершена': 50,
                'Отказ': 20,
            },
            'deal_stages': {
                'Новая': 50,
                'В работе': 40,
                'Завершена': 35,
            },
            'conversion_by_manager': {
                'Елена Юматова': {'leads': 125, 'deals': 62, 'conversion': 49.6},
                'Менеджер 2': {'leads': 125, 'deals': 63, 'conversion': 50.4},
            },
            'lead_refusals': {
                'Не дозвонился': 10,
                'Не заинтересован': 5,
                'Перезвонить позже': 5,
            },
            'deal_refusals': {
                'Передумал': 3,
                'Дорого': 2,
            },
        }
        
        # Реалистичные данные лидов
        leads_df = pd.DataFrame({
            'ID': list(range(1, 251)),
            'Название лида': [f'Лид {i}' for i in range(1, 251)],
            'Категория': ['Автосервис'] * 100 + ['Шиномонтаж'] * 80 + ['Детейлинг'] * 70,
            'Стадия': ['Новая'] * 100 + ['В работе'] * 80 + ['Завершена'] * 50 + ['Отказ'] * 20,
            'Ответственный': ['Елена Юматова'] * 125 + ['Менеджер 2'] * 125,
            'Рабочий телефон': [f'+7000000{i:03d}' for i in range(1, 251)],
            'Мобильный телефон': [''] * 250,
            'Источник телефона': ['Яндекс'] * 150 + ['2GIS'] * 100,
            'Причина отказа': [''] * 230 + ['Не дозвонился'] * 10 + ['Не заинтересован'] * 5 + ['Перезвонить позже'] * 5,
            'Дата создания': ['2024-01-01'] * 250,
            'Наш лид': [True] * 250,
        })
        
        # Реалистичные данные сделок
        deals_df = pd.DataFrame({
            'ID': list(range(1, 126)),
            'Название сделки': [f'Сделка {i}' for i in range(1, 126)],
            'Категория': ['Автосервис'] * 50 + ['Шиномонтаж'] * 40 + ['Детейлинг'] * 35,
            'Стадия сделки': ['Новая'] * 50 + ['В работе'] * 40 + ['Завершена'] * 35,
            'Ответственный': ['Елена Юматова'] * 62 + ['Менеджер 2'] * 63,
            'Сумма': [100000] * 125,
            'Валюта': ['RUB'] * 125,
            'Источник телефона': ['Яндекс'] * 75 + ['2GIS'] * 50,
            'Причина отказа': [''] * 120 + ['Передумал'] * 3 + ['Дорого'] * 2,
            'Дата создания': ['2024-01-01'] * 125,
            'Наша сделка': [True] * 125,
        })
        
        # Act
        result = create_analytics_report(metrics, leads_df, deals_df, 'report.xlsx')
        
        # Assert
        self.assertTrue(result)
        mock_workbook.assert_called_once()
        mock_wb.save.assert_called_once()


if __name__ == '__main__':
    unittest.main()
